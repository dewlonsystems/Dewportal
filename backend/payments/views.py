# =============================================================================
# DEWPORTAL BACKEND - PAYMENTS VIEWS
# =============================================================================
import io
import logging
from decimal import Decimal

from django.http import HttpResponse
from django.template.loader import render_to_string
from django.utils import timezone
from django.views.decorators.csrf import csrf_exempt
from django.utils.decorators import method_decorator
from django.db.models import Sum
from rest_framework import generics, status, viewsets
from rest_framework.response import Response
from rest_framework.permissions import IsAuthenticated, AllowAny
from rest_framework.views import APIView
from django_filters.rest_framework import DjangoFilterBackend
from rest_framework.filters import OrderingFilter, SearchFilter
from rest_framework.renderers import JSONRenderer 

from core.permissions import IsStaffOrAdmin, IsOwnerOrAdmin
from .models import Transaction
from .serializers import (
    TransactionSerializer,
    TransactionInitiateSerializer,
    MpesaCallbackSerializer,
    PaystackWebhookSerializer,
)
from .mpesa import MpesaService
from .paystack import PaystackService

logger = logging.getLogger('payments')


# =============================================================================
# Helpers
# =============================================================================

def _push_to_websocket(group_name: str, event_type: str, data: dict):
    """
    Push a message to a Django Channels group safely.
    Swallows errors so a Channels failure never breaks a payment callback.
    """
    try:
        from channels.layers import get_channel_layer
        from asgiref.sync import async_to_sync
        channel_layer = get_channel_layer()
        async_to_sync(channel_layer.group_send)(
            group_name,
            {'type': event_type, 'data': data},
        )
    except Exception as e:
        logger.error(f"WebSocket push failed [{group_name}]: {str(e)}")


def _notify_payment_update(transaction: Transaction, result_desc: str = None):
    """
    Push payment_status_update to the user's personal WS channel
    and a transaction_update to the admin channel.
    """
    payload = {
        'transaction_id': transaction.id,
        'reference':      transaction.reference,
        'status':         transaction.status,
        'amount':         str(transaction.amount),
        'payment_method': transaction.payment_method,
        'result_desc':    result_desc or '',
    }

    # ── User channel ──────────────────────────────────────────────────────────
    _push_to_websocket(
        f'user_{transaction.user.id}',
        'payment_status_update',
        payload,
    )

    # ── Admin channel ─────────────────────────────────────────────────────────
    _push_to_websocket(
        'admin_notifications',
        'transaction_update',
        {**payload, 'user_id': transaction.user.id, 'username': transaction.user.username},
    )


def _create_audit_log(user, action_type: str, description: str,
                      transaction: Transaction = None, request=None,
                      severity: str = 'info'):
    """
    Create an audit log entry, swallowing errors so audit never breaks payments.
    """
    try:
        from audit.models import AuditLog
        ip = None
        if request:
            xff = request.META.get('HTTP_X_FORWARDED_FOR')
            ip = xff.split(',')[0].strip() if xff else request.META.get('REMOTE_ADDR')

        AuditLog.objects.create(
            user=user,
            action_type=action_type,
            severity=severity,
            category='payment',
            description=description,
            ip_address=ip,
            details={
                'transaction_id':  transaction.id        if transaction else None,
                'reference':       transaction.reference if transaction else None,
                'amount':          str(transaction.amount) if transaction else None,
                'payment_method':  transaction.payment_method if transaction else None,
            },
        )
    except Exception as e:
        logger.error(f"Audit log creation failed: {str(e)}")


# =============================================================================
# Transaction ViewSet
# =============================================================================

class TransactionViewSet(viewsets.ReadOnlyModelViewSet):
    """
    Read-only viewset.
    Admin → all transactions.
    Staff → own transactions only.
    """
    queryset = Transaction.objects.filter(is_deleted=False).select_related('user')
    permission_classes = [IsAuthenticated, IsStaffOrAdmin]
    serializer_class = TransactionSerializer
    filter_backends = [DjangoFilterBackend, OrderingFilter, SearchFilter]
    filterset_fields = ['status', 'payment_method', 'user']
    search_fields = ['reference', 'provider_reference', 'user__username', 'user__email']
    ordering_fields = ['created_at', 'amount', 'status']
    ordering = ['-created_at']

    def get_queryset(self):
        qs = super().get_queryset()
        if self.request.user.role == 'staff':
            qs = qs.filter(user=self.request.user)
        return qs


# =============================================================================
# Transaction Detail
# =============================================================================

class TransactionDetailView(generics.RetrieveAPIView):
    queryset = Transaction.objects.filter(is_deleted=False).select_related('user')
    permission_classes = [IsAuthenticated, IsOwnerOrAdmin]
    serializer_class = TransactionSerializer


# =============================================================================
# Initiate Payment
# =============================================================================

class InitiatePaymentView(APIView):
    """
    POST /api/payments/initiate/
    Initiate an Mpesa STK Push or a Paystack checkout.
    """
    permission_classes = [IsAuthenticated]

    def post(self, request):
        serializer = TransactionInitiateSerializer(data=request.data)
        if not serializer.is_valid():
            return Response(serializer.errors, status=status.HTTP_400_BAD_REQUEST)

        user           = request.user
        payment_method = serializer.validated_data['payment_method']
        amount         = serializer.validated_data['amount']
        phone_number   = serializer.validated_data.get('phone_number')
        description    = serializer.validated_data.get('description', '')

        # ── Create pending transaction ─────────────────────────────────────────
        transaction = Transaction.objects.create(
            user=user,
            amount=amount,
            payment_method=payment_method,
            status='pending',
            mpesa_phone_number=phone_number if payment_method == 'mpesa' else None,
            description=description,
        )
        logger.info(f"Transaction created: {transaction.reference} | {user.username} | {amount} KES")

        _create_audit_log(
            user=user,
            action_type='transaction_initiated',
            description=f"{payment_method.title()} payment initiated: {transaction.reference}",
            transaction=transaction,
            request=request,
        )

        if payment_method == 'mpesa':
            return self._initiate_mpesa(transaction, phone_number, amount, request)
        return self._initiate_paystack(transaction, user.email, amount, request)

    # ── Mpesa ──────────────────────────────────────────────────────────────────

    def _initiate_mpesa(self, transaction, phone_number, amount, request):
        try:
            result = MpesaService().initiate_stk_push(
                phone_number=phone_number,
                amount=amount,
                transaction_reference=transaction.reference,
                account_reference=transaction.reference,
            )

            if result.get('success'):
                transaction.mpesa_checkout_request_id   = result.get('checkout_request_id')
                transaction.provider_reference          = result.get('merchant_request_id')
                transaction.mpesa_response_description  = result.get('response_description')
                transaction.save(update_fields=[
                    'mpesa_checkout_request_id',
                    'provider_reference',
                    'mpesa_response_description',
                ])
                logger.info(f"STK Push sent: {transaction.reference}")

                return Response({
                    'success': True,
                    'transaction': TransactionSerializer(transaction).data,
                    'checkout_request_id': result.get('checkout_request_id'),
                    'message': result.get('customer_message', 'STK Push sent to your phone'),
                }, status=status.HTTP_200_OK)

            # STK Push failed immediately
            transaction.status = 'failed'
            transaction.mpesa_response_description = result.get('response_description', 'STK Push failed')
            transaction.save(update_fields=['status', 'mpesa_response_description'])

            _create_audit_log(
                user=transaction.user,
                action_type='transaction_failed',
                description=f"STK Push failed: {transaction.reference}",
                transaction=transaction,
                request=request,
                severity='error',
            )

            return Response({
                'success': False,
                'error': result.get('response_description', 'STK Push initiation failed'),
            }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            logger.error(f"Mpesa initiation error [{transaction.reference}]: {str(e)}")
            transaction.status = 'failed'
            transaction.mpesa_response_description = str(e)
            transaction.save(update_fields=['status', 'mpesa_response_description'])
            return Response({'success': False, 'error': 'Payment initiation failed'},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)

    # ── Paystack ───────────────────────────────────────────────────────────────

    def _initiate_paystack(self, transaction, email, amount, request):
        try:
            result = PaystackService().initialize_transaction(
                email=email,
                amount=amount,
                reference=transaction.reference,
                metadata={
                    'user_id':        transaction.user.id,
                    'transaction_id': transaction.id,
                    'username':       transaction.user.username,
                },
            )

            if result.get('success'):
                transaction.paystack_access_code       = result.get('access_code')
                transaction.paystack_authorization_url = result.get('authorization_url')
                transaction.provider_reference         = result.get('reference')
                transaction.save(update_fields=[
                    'paystack_access_code',
                    'paystack_authorization_url',
                    'provider_reference',
                ])
                logger.info(f"Paystack initialized: {transaction.reference}")

                return Response({
                    'success': True,
                    'transaction':      TransactionSerializer(transaction).data,
                    'authorization_url': result.get('authorization_url'),
                    'access_code':      result.get('access_code'),
                    'message':          'Redirect to Paystack checkout to complete payment',
                }, status=status.HTTP_200_OK)

            transaction.status = 'failed'
            transaction.save(update_fields=['status'])

            _create_audit_log(
                user=transaction.user,
                action_type='transaction_failed',
                description=f"Paystack init failed: {transaction.reference}",
                transaction=transaction,
                request=request,
                severity='error',
            )

            return Response({
                'success': False,
                'error': result.get('error', 'Payment initialization failed'),
            }, status=status.HTTP_400_BAD_REQUEST)

        except Exception as e:
            logger.error(f"Paystack initiation error [{transaction.reference}]: {str(e)}")
            transaction.status = 'failed'
            transaction.save(update_fields=['status'])
            return Response({'success': False, 'error': 'Payment initiation failed'},
                            status=status.HTTP_500_INTERNAL_SERVER_ERROR)


# =============================================================================
# Paystack Verify  (called from frontend /payments/verify?reference=xxx)
# =============================================================================

class PaystackVerifyView(APIView):
    """
    GET /api/v1/payments/paystack/verify/?reference=DP...
    Called by the frontend verify page after Paystack redirects the user back.

    🔐 SECURITY:
    - Allows unauthenticated access (user may not be logged in after Paystack redirect)
    - Validates ownership ONLY if user is authenticated
    - Returns minimal data for unauthenticated requests
    """
    permission_classes = [AllowAny]

    def get(self, request):
        reference = request.query_params.get('reference')
        if not reference:
            return Response({'error': 'Reference is required'},
                            status=status.HTTP_400_BAD_REQUEST)

        # ── Find the transaction ───────────────────────────────────────────────
        try:
            transaction = Transaction.objects.get(reference=reference)
        except Transaction.DoesNotExist:
            return Response({'error': 'Transaction not found'},
                            status=status.HTTP_404_NOT_FOUND)

        # ── Ownership check (ONLY if user is authenticated) ────────────────────
        if request.user.is_authenticated:
            if request.user.role != 'admin' and transaction.user != request.user:
                logger.warning(
                    f"Unauthorized access attempt to transaction {reference} | "
                    f"User: {request.user.username} | IP: {request.META.get('REMOTE_ADDR')}"
                )
                return Response({'error': 'Permission denied'},
                                status=status.HTTP_403_FORBIDDEN)

        # ── Already resolved — return immediately ──────────────────────────────
        if transaction.status in ('completed', 'failed', 'cancelled'):
            if request.user.is_authenticated:
                transaction_data = TransactionSerializer(transaction).data
            else:
                transaction_data = {
                    'reference':      transaction.reference,
                    'status':         transaction.status,
                    'amount':         str(transaction.amount),
                    'payment_method': transaction.payment_method,
                    'created_at':     transaction.created_at.isoformat(),
                }

            return Response({
                'success':     transaction.status == 'completed',
                'status':      transaction.status,
                'transaction': transaction_data,
            })

        # ── Ask Paystack ───────────────────────────────────────────────────────
        try:
            result = PaystackService().verify_transaction(reference)
        except Exception as e:
            logger.error(f"Paystack verify API error [{reference}]: {str(e)}")
            return Response({'error': 'Verification service unavailable'},
                            status=status.HTTP_503_SERVICE_UNAVAILABLE)

        if not result.get('success'):
            return Response({
                'success': False,
                'status':  'pending',
                'message': result.get('error', 'Verification failed — transaction may still be processing'),
                'transaction': {
                    'reference': transaction.reference,
                    'status':    transaction.status,
                } if not request.user.is_authenticated else TransactionSerializer(transaction).data,
            })

        paystack_status = result.get('status')

        # ── Map Paystack status → our status ───────────────────────────────────
        if paystack_status == 'success':
            if transaction.can_transition_to('completed'):
                try:
                    transaction.update_status(
                        'completed',
                        callback_payload=result,
                        provider_reference=reference,
                    )
                    _notify_payment_update(transaction)

                    if request.user.is_authenticated:
                        _create_audit_log(
                            user=request.user,
                            action_type='transaction_completed',
                            description=f"Paystack payment verified & completed: {reference}",
                            transaction=transaction,
                            request=request,
                        )
                except ValueError as e:
                    logger.warning(f"Status transition error [{reference}]: {str(e)}")

            return Response({
                'success': True,
                'status':  'completed',
                'transaction': TransactionSerializer(transaction).data
                    if request.user.is_authenticated
                    else {
                        'reference': transaction.reference,
                        'status':    'completed',
                        'amount':    str(transaction.amount),
                    },
            })

        elif paystack_status in ('failed', 'abandoned'):
            if transaction.can_transition_to('failed'):
                try:
                    transaction.update_status('failed', callback_payload=result)
                    _notify_payment_update(transaction, result_desc=f"Paystack: {paystack_status}")
                except ValueError:
                    pass

            return Response({
                'success': False,
                'status':  'failed',
                'message': f'Payment {paystack_status} on Paystack',
                'transaction': TransactionSerializer(transaction).data
                    if request.user.is_authenticated
                    else {'reference': transaction.reference, 'status': 'failed'},
            })

        return Response({
            'success': False,
            'status':  'pending',
            'message': 'Payment is still being processed',
            'transaction': {
                'reference': transaction.reference,
                'status':    'pending',
            } if not request.user.is_authenticated else TransactionSerializer(transaction).data,
        })


# =============================================================================
# Mpesa Callback  (called by Safaricom Daraja — public endpoint)
# =============================================================================

@method_decorator(csrf_exempt, name='dispatch')
class MpesaCallbackView(APIView):
    """
    POST /api/payments/mpesa/callback/
    Receives STK Push result from Safaricom. Always returns 200 to prevent retries.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        serializer = MpesaCallbackSerializer(data=request.data)
        if not serializer.is_valid():
            logger.warning(f"Invalid Mpesa callback: {request.data}")
            return Response({'status': 'rejected'}, status=status.HTTP_400_BAD_REQUEST)

        try:
            callback    = MpesaService().parse_callback(request.data)
            checkout_id = callback.get('checkout_request_id')
            is_success  = callback.get('success')
            result_desc = callback.get('result_description', '')
            tx_data     = callback.get('transaction_data', {})

            try:
                transaction = Transaction.objects.select_related('user').get(
                    mpesa_checkout_request_id=checkout_id
                )
            except Transaction.DoesNotExist:
                logger.warning(f"No transaction for CheckoutRequestID: {checkout_id}")
                return Response({'status': 'accepted'})

            new_status = 'completed' if is_success else 'failed'

            if not transaction.can_transition_to(new_status):
                logger.info(f"Skipping duplicate callback for {transaction.reference} "
                            f"(already {transaction.status})")
                return Response({'status': 'accepted'})

            # ── Update transaction ─────────────────────────────────────────────
            transaction.update_status(
                new_status,
                callback_payload=request.data,
                provider_reference=tx_data.get('receipt_number') if is_success else None,
            )

            if is_success:
                transaction.mpesa_receipt_number = tx_data.get('receipt_number')
                transaction.save(update_fields=['mpesa_receipt_number'])
            else:
                transaction.mpesa_response_description = result_desc
                transaction.save(update_fields=['mpesa_response_description'])

            _notify_payment_update(
                transaction,
                result_desc=result_desc if not is_success else None,
            )

            _create_audit_log(
                user=transaction.user,
                action_type='transaction_completed' if is_success else 'transaction_failed',
                description=(
                    f"Mpesa callback received: {transaction.reference} → {new_status}"
                    + (f" | {result_desc}" if result_desc else '')
                ),
                transaction=transaction,
                severity='info' if is_success else 'warning',
            )

            logger.info(
                f"Mpesa callback processed: {transaction.reference} → {new_status}"
                + (f" ({result_desc})" if not is_success else '')
            )

        except Exception as e:
            logger.error(f"Mpesa callback error: {str(e)}", exc_info=True)

        return Response({'status': 'accepted'})


# =============================================================================
# Paystack Webhook  (called by Paystack — public endpoint)
# =============================================================================

@method_decorator(csrf_exempt, name='dispatch')
class PaystackWebhookView(APIView):
    """
    POST /api/payments/paystack/webhook/
    Receives charge.success events from Paystack.
    """
    permission_classes = [AllowAny]

    def post(self, request):
        # ── Signature verification ─────────────────────────────────────────────
        signature = request.headers.get('X-Paystack-Signature')
        if not PaystackService().verify_webhook_signature(request.body, signature):
            logger.warning("Invalid Paystack webhook signature — rejected")
            return Response({'status': 'rejected'}, status=status.HTTP_401_UNAUTHORIZED)

        try:
            data      = PaystackService().parse_webhook(request.data)
            event     = data.get('event')
            reference = data.get('reference')

            if event != 'charge.success':
                logger.info(f"Paystack webhook ignored: {event}")
                return Response({'status': 'accepted'})

            try:
                transaction = Transaction.objects.select_related('user').get(
                    reference=reference
                )
            except Transaction.DoesNotExist:
                logger.warning(f"No transaction for Paystack reference: {reference}")
                return Response({'status': 'accepted'})

            # ── Amount guard ───────────────────────────────────────────────────
            webhook_amount = data.get('amount', 0)
            if abs(float(webhook_amount) - float(transaction.amount)) > 0.01:
                logger.warning(
                    f"Amount mismatch for {reference}: "
                    f"expected {transaction.amount}, got {webhook_amount}"
                )
                return Response({'status': 'accepted'})

            if not transaction.can_transition_to('completed'):
                logger.info(f"Skipping duplicate Paystack webhook for {reference} "
                            f"(already {transaction.status})")
                return Response({'status': 'accepted'})

            transaction.update_status(
                'completed',
                callback_payload=request.data,
                provider_reference=reference,
            )

            _notify_payment_update(transaction)

            _create_audit_log(
                user=transaction.user,
                action_type='transaction_completed',
                description=f"Paystack webhook confirmed: {reference}",
                transaction=transaction,
                severity='info',
            )

            logger.info(f"Paystack webhook processed: {reference} → completed")

        except Exception as e:
            logger.error(f"Paystack webhook error: {str(e)}", exc_info=True)

        return Response({'status': 'accepted'})


# =============================================================================
# Transaction Summary
# =============================================================================

class TransactionSummaryView(APIView):
    """
    GET /api/payments/summary/
    Dashboard summary cards.
    """
    permission_classes = [IsAuthenticated, IsStaffOrAdmin]

    def get(self, request):
        user = request.user
        qs = Transaction.objects.filter(is_deleted=False)
        if user.role != 'admin':
            qs = qs.filter(user=user)

        total_revenue = qs.filter(status='completed').aggregate(
            total=Sum('amount')
        )['total'] or 0

        return Response({
            'total_revenue':          str(total_revenue),
            'currency':               'KES',
            'completed_transactions': qs.filter(status='completed').count(),
            'pending_transactions':   qs.filter(status='pending').count(),
            'failed_transactions':    qs.filter(status='failed').count(),
            'total_transactions':     qs.exclude(status='cancelled').count(),
        })


# =============================================================================
# Transaction Receipt PDF
# =============================================================================

class TransactionReceiptView(APIView):
    """
    GET /api/v1/payments/transactions/<id>/receipt/
    Generate and return a PDF receipt for a completed or failed transaction.

    Staff can only download their own receipts.
    Admin can download any receipt.
    """
    permission_classes = [IsAuthenticated, IsStaffOrAdmin]

    def get(self, request, pk):
        # ── Fetch transaction ──────────────────────────────────────────────────
        try:
            transaction = Transaction.objects.select_related('user').get(
                pk=pk,
                is_deleted=False,
            )
        except Transaction.DoesNotExist:
            return Response(
                {'error': 'Transaction not found'},
                status=status.HTTP_404_NOT_FOUND,
            )

        # ── Ownership check ────────────────────────────────────────────────────
        if request.user.role != 'admin' and transaction.user != request.user:
            return Response(
                {'error': 'Permission denied'},
                status=status.HTTP_403_FORBIDDEN,
            )

        # ── Only allow receipts for resolved transactions ──────────────────────
        if transaction.status == 'pending':
            return Response(
                {'error': 'Receipt is not available for pending transactions.'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Render HTML template ───────────────────────────────────────────────
        try:
            html_content = render_to_string(
                'payments/receipt.html',
                {
                    'transaction':  transaction,
                    'generated_at': timezone.now(),
                    'generated_by': request.user.get_full_name() or request.user.username,
                },
            )
        except Exception as e:
            logger.error(f'Receipt template render failed [{pk}]: {e}')
            return Response(
                {'error': 'Failed to generate receipt. Please try again.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        # ── Convert to PDF via WeasyPrint ──────────────────────────────────────
        try:
            from weasyprint import HTML
            from weasyprint.text.fonts import FontConfiguration

            font_config = FontConfiguration()
            pdf_bytes = HTML(string=html_content, base_url=None).write_pdf(
                font_config=font_config,
            )
        except Exception as e:
            logger.error(f'WeasyPrint PDF generation failed [{pk}]: {e}')
            return Response(
                {'error': 'PDF generation failed. Please try again.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        # ── Audit log ──────────────────────────────────────────────────────────
        _create_audit_log(
            user=request.user,
            action_type='receipt_downloaded',
            description=f"Receipt downloaded for transaction: {transaction.reference}",
            transaction=transaction,
            request=request,
        )

        # ── Return PDF response ────────────────────────────────────────────────
        filename = f'receipt_{transaction.reference}.pdf'
        response = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = len(pdf_bytes)

        logger.info(
            f'Receipt generated: {transaction.reference} | '
            f'User: {request.user.username}'
        )
        return response


# =============================================================================
# Transaction Export (PDF + Excel)
# =============================================================================

class TransactionExportView(APIView):
    """
    GET /api/v1/payments/export/?format=pdf|excel
                               &status=completed|failed|pending|cancelled
                               &payment_method=mpesa|paystack
                               &date_from=YYYY-MM-DD
                               &date_to=YYYY-MM-DD

    Admin: exports all transactions (with optional filters).
    Staff: exports only their own transactions.
    """
    permission_classes = [IsAuthenticated, IsStaffOrAdmin]
    renderer_classes = [JSONRenderer]  

    def get(self, request):
        export_format = request.query_params.get('export_format', 'pdf').lower()  

        if export_format not in ('pdf', 'excel'):
            return Response(
                {'error': 'Invalid format. Use "pdf" or "excel".'},
                status=status.HTTP_400_BAD_REQUEST,
            )

        # ── Build queryset ─────────────────────────────────────────────────────
        qs = Transaction.objects.filter(
            is_deleted=False
        ).select_related('user').order_by('-created_at')

        if request.user.role != 'admin':
            qs = qs.filter(user=request.user)

        # Optional filters
        status_filter = request.query_params.get('status')
        method_filter = request.query_params.get('payment_method')
        date_from     = request.query_params.get('date_from')
        date_to       = request.query_params.get('date_to')

        if status_filter:
            qs = qs.filter(status=status_filter)
        if method_filter:
            qs = qs.filter(payment_method=method_filter)
        if date_from:
            qs = qs.filter(created_at__date__gte=date_from)
        if date_to:
            qs = qs.filter(created_at__date__lte=date_to)

        transactions = list(qs)

        # ── Aggregate stats ────────────────────────────────────────────────────
        total_amount    = qs.filter(status='completed').aggregate(t=Sum('amount'))['t'] or Decimal('0.00')
        total_count     = qs.count()
        completed_count = qs.filter(status='completed').count()
        failed_count    = qs.filter(status='failed').count()
        pending_count   = qs.filter(status='pending').count()

        context = {
            'transactions':    transactions,
            'generated_at':    timezone.now(),
            'exported_by':     request.user.get_full_name() or request.user.username,
            'total_count':     total_count,
            'total_amount':    total_amount,
            'completed_count': completed_count,
            'failed_count':    failed_count,
            'pending_count':   pending_count,
            'status_filter':   status_filter,
            'method_filter':   method_filter,
            'date_from':       date_from,
            'date_to':         date_to,
        }

        # ── Audit log ──────────────────────────────────────────────────────────
        _create_audit_log(
            user=request.user,
            action_type='transactions_exported',
            description=(
                f"Transactions exported as {export_format.upper()}: "
                f"{total_count} records | "
                f"Filters: status={status_filter}, method={method_filter}, "
                f"date_from={date_from}, date_to={date_to}"
            ),
            request=request,
        )

        if export_format == 'pdf':
            return self._export_pdf(context)
        return self._export_excel(transactions, context)

    # ── PDF export ─────────────────────────────────────────────────────────────

    def _export_pdf(self, context):
        try:
            html_content = render_to_string('payments/export_pdf.html', context)
        except Exception as e:
            logger.error(f'Export PDF template render failed: {e}')
            return Response(
                {'error': 'Failed to generate export.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        try:
            from weasyprint import HTML
            from weasyprint.text.fonts import FontConfiguration

            font_config = FontConfiguration()
            pdf_bytes = HTML(string=html_content, base_url=None).write_pdf(
                font_config=font_config,
            )
        except Exception as e:
            logger.error(f'WeasyPrint export PDF failed: {e}')
            return Response(
                {'error': 'PDF generation failed.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        timestamp = timezone.now().strftime('%Y%m%d_%H%M')
        filename  = f'transactions_export_{timestamp}.pdf'
        response  = HttpResponse(pdf_bytes, content_type='application/pdf')
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = len(pdf_bytes)

        logger.info(
            f'PDF export: {context["total_count"]} transactions | '
            f'User: {context["exported_by"]}'
        )
        return response

    # ── Excel export ───────────────────────────────────────────────────────────

    def _export_excel(self, transactions, context):
        try:
            import openpyxl
            from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
            from openpyxl.utils import get_column_letter
        except ImportError:
            return Response(
                {'error': 'Excel export unavailable. Install openpyxl.'},
                status=status.HTTP_500_INTERNAL_SERVER_ERROR,
            )

        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = 'Transactions'

        GREEN  = '1a3d2b'
        ORANGE = 'c45c1a'
        LIGHT  = 'f8faf9'
        WHITE  = 'ffffff'

        # ── Title row ──────────────────────────────────────────────────────────
        ws.merge_cells('A1:I1')
        title_cell = ws['A1']
        title_cell.value = 'DEWLON SYSTEMS — TRANSACTIONS EXPORT'
        title_cell.font      = Font(name='Calibri', bold=True, size=14, color=WHITE)
        title_cell.fill      = PatternFill('solid', fgColor=GREEN)
        title_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[1].height = 30

        # ── Subtitle row ───────────────────────────────────────────────────────
        ws.merge_cells('A2:I2')
        sub_cell = ws['A2']
        sub_cell.value = (
            f'Generated: {context["generated_at"].strftime("%d %b %Y %H:%M")} UTC  |  '
            f'Exported by: {context["exported_by"]}  |  '
            f'Total: {context["total_count"]} records  |  '
            f'Completed Amount: KES {context["total_amount"]:,.2f}'
        )
        sub_cell.font      = Font(name='Calibri', size=9, color='555555')
        sub_cell.fill      = PatternFill('solid', fgColor='f0f4f1')
        sub_cell.alignment = Alignment(horizontal='center', vertical='center')
        ws.row_dimensions[2].height = 18

        # ── Header row ─────────────────────────────────────────────────────────
        headers    = ['#', 'Reference', 'Date & Time', 'User', 'Email', 'Method', 'Provider Ref', 'Status', 'Amount (KES)']
        header_row = 3
        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font      = Font(name='Calibri', bold=True, size=9, color=WHITE)
            cell.fill      = PatternFill('solid', fgColor=GREEN)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border    = Border(bottom=Side(style='thin', color=ORANGE))
        ws.row_dimensions[header_row].height = 22

        # ── Data rows ──────────────────────────────────────────────────────────
        thin_border = Border(bottom=Side(style='thin', color='e5e7eb'))

        status_colors = {
            'completed': ('dcfce7', '15803d'),
            'failed':    ('fee2e2', 'dc2626'),
            'pending':   ('fef3c7', 'd97706'),
            'cancelled': ('f3f4f6', '6b7280'),
        }
        method_colors = {
            'mpesa':    ('dcfce7', '15803d'),
            'paystack': ('dbeafe', '1d4ed8'),
        }

        for row_idx, tx in enumerate(transactions, start=1):
            excel_row  = row_idx + header_row
            fill_color = LIGHT if row_idx % 2 == 0 else WHITE

            row_data = [
                row_idx,
                tx.reference,
                tx.created_at.strftime('%d %b %Y %H:%M'),
                tx.user.get_full_name() or tx.user.username,
                tx.user.email,
                'M-Pesa' if tx.payment_method == 'mpesa' else 'Paystack',
                tx.provider_reference or '—',
                tx.status.title(),
                float(tx.amount),
            ]

            for col_idx, value in enumerate(row_data, start=1):
                cell = ws.cell(row=excel_row, column=col_idx, value=value)
                cell.font      = Font(name='Calibri', size=9)
                cell.fill      = PatternFill('solid', fgColor=fill_color)
                cell.border    = thin_border
                cell.alignment = Alignment(vertical='center')

                if col_idx == 8:   # Status
                    sc = status_colors.get(tx.status, ('f3f4f6', '6b7280'))
                    cell.fill      = PatternFill('solid', fgColor=sc[0])
                    cell.font      = Font(name='Calibri', size=9, bold=True, color=sc[1])
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                if col_idx == 6:   # Method
                    mc = method_colors.get(tx.payment_method, ('f3f4f6', '6b7280'))
                    cell.fill      = PatternFill('solid', fgColor=mc[0])
                    cell.font      = Font(name='Calibri', size=9, bold=True, color=mc[1])
                    cell.alignment = Alignment(horizontal='center', vertical='center')

                if col_idx == 9:   # Amount
                    cell.number_format = '#,##0.00'
                    cell.alignment     = Alignment(horizontal='right', vertical='center')
                    cell.font          = Font(name='Calibri', size=9, bold=True, color=GREEN)

                if col_idx in (2, 7):   # Reference / Provider ref
                    cell.font = Font(name='Courier New', size=8.5, color='1a3d2b')

            ws.row_dimensions[excel_row].height = 18

        # ── Summary rows ───────────────────────────────────────────────────────
        summary_row = len(transactions) + header_row + 2

        ws.merge_cells(f'A{summary_row}:H{summary_row}')
        ws[f'A{summary_row}'].value     = 'SUMMARY'
        ws[f'A{summary_row}'].font      = Font(name='Calibri', bold=True, size=9, color=WHITE)
        ws[f'A{summary_row}'].fill      = PatternFill('solid', fgColor=GREEN)
        ws[f'A{summary_row}'].alignment = Alignment(horizontal='left', vertical='center')
        ws[f'I{summary_row}'].fill      = PatternFill('solid', fgColor=GREEN)
        ws.row_dimensions[summary_row].height = 18

        summary_data = [
            ('Total Records',      context['total_count']),
            ('Completed',          context['completed_count']),
            ('Failed',             context['failed_count']),
            ('Pending',            context['pending_count']),
            ('Total Amount (KES)', float(context['total_amount'])),
        ]

        for i, (label, value) in enumerate(summary_data):
            r = summary_row + 1 + i
            ws.merge_cells(f'A{r}:H{r}')
            ws[f'A{r}'].value     = label
            ws[f'A{r}'].font      = Font(name='Calibri', size=9, color='374151')
            ws[f'A{r}'].fill      = PatternFill('solid', fgColor='f0f4f1')
            ws[f'I{r}'].value     = value
            ws[f'I{r}'].font      = Font(name='Calibri', bold=True, size=9, color=GREEN)
            ws[f'I{r}'].fill      = PatternFill('solid', fgColor='f0f4f1')
            if label == 'Total Amount (KES)':
                ws[f'I{r}'].number_format = '#,##0.00'
            ws.row_dimensions[r].height = 16

        # ── Column widths ──────────────────────────────────────────────────────
        col_widths = [5, 16, 18, 22, 28, 11, 20, 12, 14]
        for i, width in enumerate(col_widths, start=1):
            ws.column_dimensions[get_column_letter(i)].width = width

        ws.freeze_panes = f'A{header_row + 1}'

        # ── Write to buffer ────────────────────────────────────────────────────
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        excel_bytes = buffer.getvalue()

        timestamp = timezone.now().strftime('%Y%m%d_%H%M')
        filename  = f'transactions_export_{timestamp}.xlsx'
        response  = HttpResponse(
            excel_bytes,
            content_type='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet',
        )
        response['Content-Disposition'] = f'attachment; filename="{filename}"'
        response['Content-Length'] = len(excel_bytes)

        logger.info(
            f'Excel export: {context["total_count"]} transactions | '
            f'User: {context["exported_by"]}'
        )
        return response